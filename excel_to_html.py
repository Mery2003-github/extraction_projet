import zipfile
import xml.etree.ElementTree as ET
import base64
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image
import io
from openpyxl.utils import range_boundaries

EMU_PER_PIXEL = 9525
PIXELS_PER_CHAR = 7  
LINE_HEIGHT = 19     
MAX_IMAGE_WIDTH = 300
QUALITY = 50

PIXELS_PER_POINT = 1.33  
DEFAULT_COL_WIDTH = 8.43  
DEFAULT_ROW_HEIGHT = 15  
   
def argb_to_hex(argb):
    if argb is None:
        return None
    argb = argb.upper()
    if len(argb) == 8:  # ARGB
        return argb[2:]
    elif len(argb) == 6:  # déjà RGB
        return argb
    else:
        return argb
def extract_styles_from_xml(zipf):
    styles = {}
    try:
        with zipf.open('xl/styles.xml') as f:
            styles_xml = ET.parse(f).getroot()
            ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

            fonts = {}
            for i, font in enumerate(styles_xml.findall('main:fonts/main:font', ns)):
                font_name_elem = font.find('main:name', ns)
                fonts[i] = {
                    'bold': font.find('main:b', ns) is not None,
                    'italic': font.find('main:i', ns) is not None,
                    'underline': font.find('main:u', ns) is not None,
                    'size': float(font.find('main:sz', ns).get('val')) if font.find('main:sz', ns) is not None else 11,
                    'color': argb_to_hex(font.find('main:color', ns).get('rgb')) if font.find('main:color', ns) is not None else None,
                    'name': font_name_elem.get('val') if font_name_elem is not None else 'Calibri'
                }


            fills = {}
            for i, fill in enumerate(styles_xml.findall('main:fills/main:fill', ns)):
                    pattern = fill.find('main:patternFill', ns)
                    gradient = fill.find('main:gradientFill', ns)
                    if pattern is not None:
                      fgColor = pattern.find('main:fgColor', ns)
                      fills[i] = {
                      'bg_color': argb_to_hex(fgColor.get('rgb')) if fgColor is not None else None
                                  }
                    elif gradient is not None:
        # gérer le gradient (optionnel)
                      fills[i] = {
                        'bg_color': None  # ou gérer la couleur dominante
                 }


            borders = {}
            for i, border in enumerate(styles_xml.findall('main:borders/main:border', ns)):
                borders[i] = {}
                for side in ['left', 'right', 'top', 'bottom']:
                    side_elem = border.find(f'main:{side}', ns)
                    if side_elem is not None:
                        color_elem = side_elem.find('main:color', ns)
                        borders[i][side] = {
                            'style': side_elem.get('style'),
                            'color': argb_to_hex(color_elem.get('rgb')) if color_elem is not None else '000000'
                        }

            alignments = {}
            for i, xf in enumerate(styles_xml.findall('main:cellXfs/main:xf', ns)):
                align = xf.find('main:alignment', ns)
                alignments[i] = {
                    'fontId': int(xf.get('fontId', 0)),
                    'fillId': int(xf.get('fillId', 0)),
                    'borderId': int(xf.get('borderId', 0)),
                    'wrap': align.get('wrapText') == '1' if align is not None else False,
                    'horizontal': align.get('horizontal') if align is not None else 'general',
                    'vertical': align.get('vertical') if align is not None else 'bottom'
                }

            styles = {
                'fonts': fonts,
                'fills': fills,
                'borders': borders,
                'alignments': alignments
            }

    except Exception as e:
        print(f"⚠ Erreur lors de la lecture du fichier styles.xml: {e}")

    return styles

def get_cell_style(cell, styles, ws=None):
    style = {
        'bold': False,
        'italic': False,
        'underline': False,
        'size': 11,
        'color': None,
        'bg_color': None,
        'border': {},
        'wrap': False,
        'align': 'general',
        'font': 'Calibri'
    }

    if not styles:
        return style

    style_index = getattr(cell, 'style_id', 0)
    alignment = styles['alignments'].get(style_index, {})
    font = styles['fonts'].get(alignment.get('fontId', 0), {})
    fill = styles['fills'].get(alignment.get('fillId', 0), {})
    border = styles['borders'].get(alignment.get('borderId', 0), {})

    style.update({
        'bold': font.get('bold', False),
        'italic': font.get('italic', False),
        'underline': font.get('underline', False),
        'size': font.get('size', 11),
        'color': font.get('color'),
        'bg_color': fill.get('bg_color'),
        'border': border,
        'wrap': alignment.get('wrap', False),
        'align': alignment.get('horizontal', 'general'),
        'vertical': alignment.get('vertical', 'bottom'),
        'font': font.get('name', 'Calibri')
    })

   
    if style['bg_color'] is None and ws is not None:
        col_letter = get_column_letter(cell.column)
        col_fill = ws.column_dimensions[col_letter].fill
        if col_fill and hasattr(col_fill, 'fgColor') and col_fill.fgColor.rgb:
            style['bg_color'] = col_fill.fgColor.rgb[2:]  # retirer AR

        row_fill = ws.row_dimensions[cell.row].fill
        if style['bg_color'] is None and row_fill and hasattr(row_fill, 'fgColor') and row_fill.fgColor.rgb:
            style['bg_color'] = row_fill.fgColor.rgb[2:]

    return style

def get_sheet_data(wb, zipf, sheet_name=None):
    styles = extract_styles_from_xml(zipf)
    ws = wb[sheet_name] if sheet_name else wb.active
    data = []

    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            value = cell.value if cell.value is not None else ""
            style = get_cell_style(cell, styles)
            row_data.append({
                'value': str(value),
                'style': style,
                'row': cell.row,
                'col': cell.column
            })
        data.append(row_data)

    return data

def get_image_data(zipf, media_path):
    try:
        with zipf.open(media_path) as f:
            data = f.read()

        img = Image.open(io.BytesIO(data))

        if img.mode in ("P", "LA"):
            img = img.convert("RGBA")
        elif img.mode != "RGB":
            img = img.convert("RGB")

      
        if img.width > MAX_IMAGE_WIDTH:
            ratio = MAX_IMAGE_WIDTH / img.width
            new_height = int(img.height * ratio)
            img = img.resize((MAX_IMAGE_WIDTH, new_height), Image.LANCZOS)

        output = io.BytesIO()
        img.save(output, format="WEBP", quality=QUALITY, method=6)
        optimized_data = output.getvalue()
        return f"data:image/webp;base64,{base64.b64encode(optimized_data).decode()}"

    except Exception as e:
        print(f"⚠ Erreur lors du traitement de l'image {media_path}: {e}")
        try:
            ext = media_path.split('.')[-1].lower()
            mime_fallback = "image/png" if ext == "png" else "image/jpeg"
            return f"data:{mime_fallback};base64," + base64.b64encode(data).decode()
        except:
            return ""  

def column_width_to_pixels(width):
    if width is None:
        width = DEFAULT_COL_WIDTH
    if width <= 0:
        return 0
    pixels = int(round(width * 5 )) 
    return pixels
def row_height_to_pixels(height):
    if height is None:
        height = DEFAULT_ROW_HEIGHT
    return int(round(height * 96 / 89))  


def get_column_widths(wb, sheet_name=None):
    ws = wb[sheet_name] if sheet_name else wb.active
    widths = {}
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = ws.column_dimensions[letter].width if letter in ws.column_dimensions else DEFAULT_COL_WIDTH
        widths[col] = column_width_to_pixels(width)
    return widths



def get_row_heights(wb, sheet_name=None):
    ws = wb[sheet_name] if sheet_name else wb.active
    heights = {}
    for row in range(1, ws.max_row + 1):
        height = ws.row_dimensions[row].height if row in ws.row_dimensions else DEFAULT_ROW_HEIGHT
        heights[row] = row_height_to_pixels(height)
    return heights


def calculate_position(start_idx, offset_emu, dimensions, is_column=True):
   
    if is_column:
        
        total = sum(dimensions.get(i, DEFAULT_COL_WIDTH) * PIXELS_PER_POINT 
                   for i in range(1, start_idx))
    else:
       
        total = sum(dimensions.get(i, DEFAULT_ROW_HEIGHT) * PIXELS_PER_POINT 
                   for i in range(1, start_idx))
    return total + (offset_emu / EMU_PER_PIXEL)

def parse_drawing(zipf, drawing_path, col_widths, row_heights):
    try:
        rels_path = drawing_path.replace('drawings/', 'drawings/_rels/') + '.rels'
        with zipf.open(rels_path) as f:
            rels_root = ET.parse(f).getroot()

        rels = {rel.attrib['Id']: rel.attrib['Target'].replace('../', 'xl/')
                for rel in rels_root.findall('{*}Relationship')}

        ns = {
            'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
            'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
            'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
        }

        with zipf.open(drawing_path) as f:
            root = ET.parse(f).getroot()

        images = []
        anchors = root.findall('xdr:oneCellAnchor', ns) + root.findall('xdr:twoCellAnchor', ns)

        for anchor in anchors:
            from_elem = anchor.find('xdr:from', ns)
            to_elem = anchor.find('xdr:to', ns)

            if from_elem is None:
                continue

            col = int(from_elem.find('xdr:col', ns).text) + 1
            row = int(from_elem.find('xdr:row', ns).text) + 1
            colOff = int(from_elem.find('xdr:colOff', ns).text)
            rowOff = int(from_elem.find('xdr:rowOff', ns).text)

            if to_elem is not None:
                to_col = int(to_elem.find('xdr:col', ns).text) + 1
                to_row = int(to_elem.find('xdr:row', ns).text) + 1
                to_colOff = int(to_elem.find('xdr:colOff', ns).text)
                to_rowOff = int(to_elem.find('xdr:rowOff', ns).text)

                left = calculate_position(col, colOff, col_widths, is_column=True)
                top = calculate_position(row, rowOff, row_heights, is_column=False)
                right = calculate_position(to_col, to_colOff, col_widths, is_column=True)
                bottom = calculate_position(to_row, to_rowOff, row_heights, is_column=False)

                width_px = right - left
                height_px = bottom - top
            else:
                ext = anchor.find('xdr:ext', ns)
                if ext is None:
                    continue
                cx, cy = int(ext.attrib['cx']), int(ext.attrib['cy'])
                width_px = cx / EMU_PER_PIXEL
                height_px = cy / EMU_PER_PIXEL

                left = calculate_position(col, colOff, col_widths, is_column=True)
                top = calculate_position(row, rowOff, row_heights, is_column=False)

            blip = anchor.find('.//a:blip', ns)
            if blip is None:
                continue
            embed = blip.attrib.get(f"{{{ns['r']}}}embed")
            if embed not in rels:
                continue

            images.append({
                'row': row,
                'col': col,
                'left': left,
                'top': top,
                'width': width_px,
                'height': height_px,
                'data_uri': get_image_data(zipf, rels[embed]),
                'cell_width': col_widths.get(col, DEFAULT_COL_WIDTH) * PIXELS_PER_POINT,
                'cell_height': row_heights.get(row, DEFAULT_ROW_HEIGHT) * PIXELS_PER_POINT
            })

        return images

    except Exception as e:
        print(f"⚠ Erreur lors de l'analyse du dessin: {e}")
        return []

def estimate_text_height(text, font_size=11, cell_width_px=100, wrap=True):
    if not wrap:
        return font_size * 1.2  # Une seule ligne

    lines = text.split('\n')
    avg_char_width = font_size * 0.6
    chars_per_line = max(1, int(cell_width_px / avg_char_width))
    total_lines = 0
    for line in lines:
        line_len = len(line)
        wrapped_lines = max(1, (line_len + chars_per_line - 1) // chars_per_line)
        total_lines += wrapped_lines

    return total_lines * font_size * 1.2

def points_to_pixels(pt):
    return round(pt * 1.3333)
from PIL import ImageFont

import xml.etree.ElementTree as ET

def get_sheet_zoom(zipf, sheet_path):
    try:
        with zipf.open(sheet_path) as f:
            root = ET.parse(f).getroot()
            ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            sheetViews = root.find('main:sheetViews', ns)
            if sheetViews is not None:
                sheetView = sheetViews.find('main:sheetView', ns)
                if sheetView is not None and 'zoomScale' in sheetView.attrib:
                    zoom_scale = int(sheetView.attrib['zoomScale'])
                    return zoom_scale
    except Exception as e:
        print(f"⚠ Erreur en lisant le zoom de la feuille: {e}")
    return 100  

def get_text_size(text, font_path, font_size, max_width=None):
    font = ImageFont.truetype(font_path, font_size)
    lines = text.split('\n')
    max_line_width = 0
    total_height = 0
    line_spacing = int(font_size * 0.2)

    for line in lines:
        if max_width:
            words = line.split(' ')
            current_line = ''
            wrapped_lines = []
            for word in words:
                test_line = current_line + (' ' if current_line else '') + word
                w, h = font.getsize(test_line)
                if w > max_width and current_line:
                    wrapped_lines.append(current_line)
                    current_line = word
                else:
                    current_line = test_line
            wrapped_lines.append(current_line)
        else:
            wrapped_lines = [line]

        for wrapped_line in wrapped_lines:
            w, h = font.getsize(wrapped_line)
            if w > max_line_width:
                max_line_width = w
            total_height += h + line_spacing
    return max_line_width, total_height
def points_to_pixels(pt):
    return round(pt * 1.3333)
def get_border_css_full(style):
    if style is None or style.get("style") is None:
        return ""

    style_map = {
        "hair": "0.5px solid",
        "thin": "1px solid",
        "medium": "2px solid",
        "thick": "3px solid",
        "dashed": "1px dashed",
        "dotted": "1px dotted",
        "double": "3px double"
    }

    border_style = style_map.get(style.get("style"), "1px solid")
    color = f"#{style.get('color')}" if style.get("color") else "black"
    return f"{border_style} {color}"


def border_to_style_full(cell_style):
    css = ""
    borders = cell_style.get("border", {})

    for side in ['top', 'bottom', 'left', 'right']:
        if side in borders and borders[side].get("style"):
            border_css = get_border_css_full(borders[side])
            css += f"border-{side}:{border_css};"

    return css
from collections import defaultdict

from collections import defaultdict
def print_dimensions_before_after(col_widths, row_heights, target_width, tolerance=0.01):
    
    original_width = sum(col_widths.values()) * PIXELS_PER_POINT
    original_height = sum(h * PIXELS_PER_POINT for h in row_heights.values())
    aspect_ratio_original = original_width / original_height

    new_width = target_width
    new_height = int(target_width / aspect_ratio_original)
    aspect_ratio_new = new_width / new_height

    print(" Dimensions de la feuille Excel :")
    print(f"  - Largeur originale : {original_width:.2f}px")
    print(f"  - Hauteur originale : {original_height:.2f}px")
    print(f"  - Ratio d'aspect original (W/H) : {aspect_ratio_original:.4f}")
    print()
    print(" Dimensions dans le HTML exporté :")
    print(f"  - Largeur cible : {new_width}px")
    print(f"  - Hauteur calculée : {new_height}px")
    print(f"  - Ratio d'aspect nouveau (W/H) : {aspect_ratio_new:.4f}")
    print()

 
    if abs(aspect_ratio_original - aspect_ratio_new) <= tolerance:
        print(" Le ratio d'aspect est respecté (avant et après extraction).")
    else:
        print("⚠ Le ratio d'aspect **n'est pas** respecté (avant et après extraction).")
    print("---------------------------------------------------")

def generate_html(sheet_data, images, col_widths, row_heights, output_file, zoom_scale=100, target_width=550):
    print_dimensions_before_after(col_widths, row_heights, target_width)

    original_width = sum(col_widths.values()) * PIXELS_PER_POINT
    original_height = sum(h * PIXELS_PER_POINT for h in row_heights.values())
    aspect_ratio = original_width / original_height

    new_width = target_width
    new_height = int(target_width / aspect_ratio)
    scale_x = new_width / original_width
    scale_y = new_height / original_height

    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Export Excel fidèle</title>
</head>
<body>
<div style="position:relative;width:{new_width}px;height:{new_height}px;">
"""

   
    for row in sheet_data:
        for cell in row:
            left = calculate_position(cell['col'], 0, col_widths, True) * scale_x
            top = calculate_position(cell['row'], 0, row_heights, False) * scale_y
            width = col_widths.get(cell['col'], DEFAULT_COL_WIDTH) * PIXELS_PER_POINT * scale_x
            height = row_heights.get(cell['row'], DEFAULT_ROW_HEIGHT) * PIXELS_PER_POINT * scale_y

            font_px = points_to_pixels(cell['style']['size']) * (zoom_scale / 100) * scale_y

            vertical_align = cell['style'].get('vertical', 'bottom')
            cell_height = row_heights.get(cell['row'], DEFAULT_ROW_HEIGHT) * PIXELS_PER_POINT * scale_y

            if vertical_align == 'center':
                top += (cell_height - height) / 2
            elif vertical_align == 'bottom':
                top += cell_height - height

            style = f"font-size:{font_px}px; font-family:'{cell['style'].get('font', 'Calibri')}', sans-serif;"
            if cell['style']['bold']:
                style += "font-weight:bold;"
            if cell['style']['italic']:
                style += "font-style:italic;"
            if cell['style']['underline']:
                style += "text-decoration:underline;"
            if cell['style']['color']:
                style += f"color:#{cell['style']['color']};"
            if cell['style']['bg_color']:
                style += f"background-color:#{cell['style']['bg_color']};"

            style += border_to_style_full(cell['style'])

            align_map = {
                'left': 'left',
                'right': 'right',
                'center': 'center',
                'justify': 'justify',
                'general': 'left',
                'distributed': 'justify'
            }
            align = align_map.get(cell['style']['align'], 'left')
            style += f"text-align:{align};"

            if cell['style']['wrap']:
                style += "white-space:normal; overflow:visible;"
            else:
                style += "white-space:nowrap;"

            cell_value_html = cell['value'].replace('\n', '<br>')

            html += f"""
<div style="position:absolute; left:{left}px; top:{top}px; width:{width}px; height:{height}px; {style} box-sizing:border-box; overflow:hidden;">
    {cell_value_html}
</div>
"""


    images_by_cell = defaultdict(list)
    for img in images:
        images_by_cell[(img['row'], img['col'])].append(img)

  
    for (row, col), cell_images in images_by_cell.items():
        cell_left = calculate_position(col, 0, col_widths, True) * scale_x
        cell_top = calculate_position(row, 0, row_heights, False) * scale_y
        cell_width = col_widths.get(col, DEFAULT_COL_WIDTH) * PIXELS_PER_POINT * scale_x
        cell_height = row_heights.get(row, DEFAULT_ROW_HEIGHT) * PIXELS_PER_POINT * scale_y

        html += f"""
<div style="position:absolute; left:{cell_left}px; top:{cell_top}px; width:{cell_width}px; height:{cell_height}px;
            display:flex; flex-wrap:wrap; align-items:center; justify-content:center; box-sizing:border-box;">
"""

        for img in cell_images:
            img_width = img['width'] * scale_x
            img_height = img['height'] * scale_y
            html += f"""
    <img src="{img['data_uri']}" alt="Image" style="max-width:{img_width}px; max-height:{img_height}px;
             object-fit:contain; margin:2px;">
"""

        html += "</div>"

    html += """
</div>
</body>
</html>
"""

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"✅ Fichier HTML généré avec échelle : {output_file}")

def main():
    input_file = r"xlsx\Etiquette CLEMENTINE (10).xlsx"
    output_file = "fidele.html"


    wb = load_workbook(input_file)
    sheet_name = wb.sheetnames[0]

    with zipfile.ZipFile(input_file) as zipf:
      
        sheet_data = get_sheet_data(wb, zipf, sheet_name)
        col_widths = get_column_widths(wb, sheet_name)
        row_heights = get_row_heights(wb, sheet_name)

       
        sheet_path = f'xl/worksheets/{sheet_name}.xml'  
        
      
        sheet_path = [f for f in zipf.namelist() if f.startswith('xl/worksheets/sheet')][0]

        zoom_scale = get_sheet_zoom(zipf, sheet_path)
      
        all_images = []
        drawings = [f for f in zipf.namelist() if f.startswith('xl/drawings/drawing')]

        for drawing_path in drawings:
            images = parse_drawing(zipf, drawing_path, col_widths, row_heights)
            all_images.extend(images)

  
    generate_html(sheet_data, all_images, col_widths, row_heights, output_file, zoom_scale)

if __name__ == "__main__":
    main()
